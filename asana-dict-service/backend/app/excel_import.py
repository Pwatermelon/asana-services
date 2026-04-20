"""
Модуль для импорта данных из Excel файлов
"""
import base64
import json
import logging
import re
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, List, Optional, Tuple, Callable

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.config import logger, SQLALCHEMY_DATABASE_URL
from app.models import ModerationItem
from app.ontology import load_asana_names, add_asana_name, add_asana, find_existing_source, add_source
from app.s3_utils import upload_image_to_s3
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

def is_base64_string(value: str) -> bool:
    """
    Проверяет, является ли строка валидной base64 строкой (без префикса data:).
    Base64 строка должна быть достаточно длинной и содержать только допустимые символы.
    """
    if not isinstance(value, str):
        return False
    
    # Убираем пробелы и переносы строк
    value = value.strip().replace('\n', '').replace('\r', '').replace(' ', '')
    
    # Base64 строка должна быть достаточно длинной (минимум 20 символов для маленького изображения)
    if len(value) < 20:
        return False
    
    # Base64 содержит только A-Z, a-z, 0-9, +, /, = (padding)
    base64_pattern = re.compile(r'^[A-Za-z0-9+/]+=*$')
    if not base64_pattern.match(value):
        return False
    
    # Проверяем, что это действительно base64 (попробуем декодировать)
    try:
        decoded = base64.b64decode(value, validate=True)
        # Если декодирование успешно и результат не пустой - это base64
        if len(decoded) > 0:
            # Проверяем, что это похоже на изображение (проверяем магические байты)
            if (decoded.startswith(b'\x89PNG') or  # PNG
                decoded.startswith(b'\xff\xd8\xff') or  # JPEG
                decoded.startswith(b'GIF8') or  # GIF
                (decoded.startswith(b'RIFF') and b'WEBP' in decoded[:12])):  # WebP
                return True
    except Exception:
        pass
    
    return False


def diagnose_image_type(img, row: int):
    """
    Диагностирует тип изображения в Excel - встроенное или ссылка.
    """
    # Проверяем, есть ли путь к файлу (это признак ссылки)
    if hasattr(img, 'path'):
        return "link"
    
    # Проверяем, есть ли данные изображения
    has_data = False
    if hasattr(img, '_data') and img._data:
        has_data = True
    if hasattr(img, 'ref') and img.ref:
        if isinstance(img.ref, bytes):
            has_data = True
        elif hasattr(img.ref, 'read'):
            has_data = True
    
    if has_data:
        return "embedded"
    else:
        return "link"


def extract_images_from_worksheet(ws) -> Dict[int, List[str]]:
    """
    Извлекает все изображения из листа Excel и возвращает словарь {row_number: [base64, ...]}.
    Привязывает изображения к строке, в которой они находятся.
    Поддерживает несколько изображений в одной строке.
    Использует упрощенную логику извлечения через img._data().
    """
    images = {}
    try:
        # Проверяем наличие изображений в листе
        if not hasattr(ws, '_images'):
            return images
            
        if not ws._images:
            return images
        
        for img_idx, img in enumerate(ws._images):
            try:
                # Получаем координаты ячейки (используем логику из рабочего скрипта)
                row = None
                
                if hasattr(img, 'anchor') and img.anchor:
                    if hasattr(img.anchor, '_from'):
                        if hasattr(img.anchor._from, 'row'):
                            row = img.anchor._from.row + 1  # +1 потому что openpyxl индексирует с 0
                    elif hasattr(img.anchor, 'min_row'):
                        row = img.anchor.min_row
                    elif hasattr(img.anchor, 'row'):
                        row = img.anchor.row + 1
                
                if not row or row == 1:
                    continue
                
                # Получаем данные изображения через _data() (как в рабочем скрипте)
                img_data = None
                try:
                    if hasattr(img, '_data'):
                        # Пробуем вызвать как функцию
                        if callable(img._data):
                            img_data = img._data()
                        else:
                            img_data = img._data
                except Exception as e:
                    logger.warning(f"Строка {row}: не удалось извлечь данные изображения через _data(): {e}")
                    continue
                
                if not img_data:
                    logger.warning(f"Строка {row}: данные изображения пустые")
                    continue
                
                # Проверяем, что данные не пустые и достаточно большие
                if len(img_data) < 100:
                    logger.warning(f"Строка {row}: данные изображения слишком маленькие ({len(img_data)} байт), пропускаем")
                    continue
                
                # Конвертируем в base64 (без префикса data:, как в рабочем скрипте)
                try:
                    base64_str = base64.b64encode(img_data).decode('utf-8')
                except Exception as e:
                    logger.error(f"Строка {row}: ошибка кодирования изображения в base64: {e}")
                    continue
                
                # Добавляем изображение в список для этой строки (поддерживаем несколько фото)
                if row not in images:
                    images[row] = []
                images[row].append(base64_str)
                
            except Exception as e:
                logger.warning(f"Строка {row if 'row' in locals() else '?'}: ошибка при обработке изображения: {e}")
                continue
        
    except Exception as e:
        logger.error(f"Ошибка при извлечении изображений из листа: {e}", exc_info=True)
    
    return images


def normalize_column_names(row: Dict[str, Any]) -> Dict[str, Any]:
    """Нормализует имена колонок в строке данных"""
    normalized = {}
    column_mapping = {
        'название': 'name_ru',
        'название асаны': 'name_ru',
        'name_ru': 'name_ru',
        'название асаны (рус)': 'name_ru',
        'название на русском': 'name_ru',
        'санскрит': 'name_sanskrit',
        'name_sanskrit': 'name_sanskrit',
        'название на санскрите': 'name_sanskrit',
        'транслитерация': 'transliteration',
        'transliteration': 'transliteration',
        'определение': 'definition',
        'definition': 'definition',
        'фото': 'photo',
        'photo': 'photo',
        'изображение': 'photo',
        'image': 'photo',
        'название источника': 'source_title',
        'source_title': 'source_title',
        'title': 'source_title',
        'автор': 'source_author',
        'source_author': 'source_author',
        'author': 'source_author',
        'год': 'source_year',
        'source_year': 'source_year',
        'year': 'source_year',
        'издательство': 'source_publisher',
        'source_publisher': 'source_publisher',
        'publisher': 'source_publisher',
        'страницы': 'source_pages',
        'source_pages': 'source_pages',
        'pages': 'source_pages',
        'аннотация': 'source_annotation',
        'source_annotation': 'source_annotation',
        'annotation': 'source_annotation'
    }
    
    for key, value in row.items():
        key_lower = key.lower().strip()
        normalized_key = column_mapping.get(key_lower, key_lower)
        normalized[normalized_key] = value
    
    return normalized


def find_matching_asana_name(name_ru: str, fuzzy_threshold: float = 1.0) -> tuple:
    """
    Ищет существующее название асаны ТОЛЬКО по точному совпадению (100%, без учета регистра).
    Возвращает (ID найденного названия, None) если найдено точное совпадение,
    или (None, None) если не найдено точного совпадения.
    """
    try:
        existing_names = load_asana_names()
        if not existing_names:
            logger.warning("No existing asana names found in ontology")
            return None, None
        
        name_ru_lower = name_ru.lower().strip()
        
        # Проверяем ТОЛЬКО точное совпадение (100%, без учета регистра)
        for name_data in existing_names:
            existing_name_lower = name_data.get("name_ru", "").lower().strip()
            if existing_name_lower == name_ru_lower:
                logger.info(f"Found exact match (100%, case-insensitive) for '{name_ru}' -> '{name_data.get('name_ru')}'")
                return name_data["id"], None
        
        # Если точного совпадения нет, возвращаем None (не ищем похожие)
        logger.info(f"No exact match (100%) found for '{name_ru}'")
        return None, None
            
    except Exception as e:
        logger.error(f"Error finding matching asana name: {e}")
        return None, None


def has_source_data(row: Dict[str, Any]) -> bool:
    """Проверяет, есть ли в строке данные для источника"""
    source_fields = ['source_title', 'source_author', 'source_year', 
                     'source_publisher', 'source_pages', 'source_annotation']
    return any(row.get(field) for field in source_fields)


def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """Парсит Excel файл и возвращает список словарей с данными"""
    try:
        # НЕ используем data_only=True, так как это может помешать загрузке изображений
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        # Извлекаем все изображения из листа (если ошибка - продолжаем без изображений)
        images = {}
        try:
            images = extract_images_from_worksheet(ws)
        except Exception as img_error:
            logger.warning(f"Ошибка при извлечении изображений из Excel (продолжаем без изображений): {img_error}")
            images = {}
        
        # Определяем заголовки (первая строка)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.lower().strip() if cell.value else '')
        
        # Парсим данные
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            photo_base64_from_cell = None
            
            for col_idx, cell in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else f'column_{col_idx}'
                
                if cell.value is not None:
                    # Проверяем, является ли ячейка с заголовком "photo" или похожим base64 строкой
                    if header and ('photo' in header.lower() or 'фото' in header.lower()):
                        cell_value = str(cell.value).strip()
                        if is_base64_string(cell_value):
                            photo_base64_from_cell = cell_value
                            # Не добавляем в row_data, обработаем отдельно
                            continue
                    
                    row_data[header] = cell.value
            
            # Приоритет: сначала base64 из ячейки, потом встроенные изображения
            if photo_base64_from_cell:
                # Используем base64 из ячейки (без префикса data:, он будет добавлен позже)
                row_data['photo'] = photo_base64_from_cell
            elif row_idx in images:
                # Используем встроенные изображения, если нет base64 в ячейке
                photos_list = images[row_idx]
                # Если одно фото, сохраняем как строку для обратной совместимости
                # Если несколько фото, сохраняем как список
                if len(photos_list) == 1:
                    row_data['photo'] = photos_list[0]
                else:
                    row_data['photos'] = photos_list
                    row_data['photo'] = photos_list[0]  # Первое фото для обратной совместимости
            
            # Пропускаем пустые строки
            if any(row_data.values()):
                rows.append(row_data)
        
        return rows
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise


def get_db_session():
    """Создает и возвращает сессию базы данных"""
    engine = create_engine(SQLALCHEMY_DATABASE_URL)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    return SessionLocal()


_RE_ROW_PREFIX = re.compile(r"^\s*Строка\s+\d+:\s*", re.IGNORECASE)


def _canonical_moderation_error(msg: Optional[str]) -> str:
    """Убирает префикс с номером строки Excel — при повторном импорте того же файла номер может отличаться."""
    if not msg:
        return ""
    return _RE_ROW_PREFIX.sub("", msg).strip().lower()


def _suggestion_key(
    suggested_name_ru: Optional[str],
    suggested_name_sanskrit: Optional[str],
    suggested_transliteration: Optional[str],
    suggested_definition: Optional[str],
    existing_name_id: Optional[str],
    existing_name_ru: Optional[str],
) -> Tuple[str, ...]:
    return (
        (suggested_name_ru or "").strip().lower(),
        (suggested_name_sanskrit or "").strip().lower(),
        (suggested_transliteration or "").strip().lower(),
        (suggested_definition or "").strip().lower()[:500] if suggested_definition else "",
        (existing_name_id or "").strip().lower(),
        (existing_name_ru or "").strip().lower(),
    )


def _suggestion_key_from_row(row: ModerationItem) -> Tuple[str, ...]:
    return _suggestion_key(
        row.suggested_name_ru,
        row.suggested_name_sanskrit,
        row.suggested_transliteration,
        row.suggested_definition,
        row.existing_name_id,
        row.existing_name_ru,
    )


def _find_duplicate_unresolved(
    db,
    asana_name: str,
    source_id: Optional[str],
    error_message: str,
    moderation_type: str,
    object_type: Optional[str],
    suggested_name_ru: Optional[str],
    suggested_name_sanskrit: Optional[str],
    suggested_transliteration: Optional[str],
    suggested_definition: Optional[str],
    existing_name_id: Optional[str],
    existing_name_ru: Optional[str],
) -> Optional[ModerationItem]:
    q = (
        db.query(ModerationItem)
        .filter(ModerationItem.resolved.is_(False))
        .filter(ModerationItem.moderation_type == moderation_type)
        .filter(ModerationItem.asana_name == (asana_name or ""))
    )
    if source_id is None:
        q = q.filter(ModerationItem.source_id.is_(None))
    else:
        q = q.filter(ModerationItem.source_id == source_id)
    if object_type is None:
        q = q.filter(ModerationItem.object_type.is_(None))
    else:
        q = q.filter(ModerationItem.object_type == object_type)

    canon_new = _canonical_moderation_error(error_message)
    key_new = _suggestion_key(
        suggested_name_ru,
        suggested_name_sanskrit,
        suggested_transliteration,
        suggested_definition,
        existing_name_id,
        existing_name_ru,
    )

    for row in q.all():
        if _canonical_moderation_error(row.error_message) != canon_new:
            continue
        if _suggestion_key_from_row(row) != key_new:
            continue
        return row
    return None


def save_moderation_item(
    asana_name: str,
    source_id: Optional[str],
    error_message: str,
    row_number: int,
    import_data: Optional[Dict[str, Any]] = None,
    user: Optional[str] = None,
    moderation_type: str = "error",
    suggested_name_ru: Optional[str] = None,
    suggested_name_sanskrit: Optional[str] = None,
    suggested_transliteration: Optional[str] = None,
    suggested_definition: Optional[str] = None,
    existing_name_id: Optional[str] = None,
    existing_name_ru: Optional[str] = None,
    object_type: Optional[str] = None  # 'asana_name', 'source', 'asana'
):
    """Сохраняет запись в таблицу модерации"""
    try:
        # Определяем тип объекта, если не указан
        if not object_type:
            if moderation_type in ['duplicate_name', 'name_mismatch'] and not source_id:
                # Если это проблема с названием и нет источника - это название асаны
                object_type = 'asana_name'
            elif moderation_type == 'duplicate_source' or (source_id and 'source' in error_message.lower()):
                # Если это дубликат источника или ошибка связана с источником - это источник
                object_type = 'source'
            elif source_id and asana_name:
                # Если есть и источник, и название асаны - это асана
                object_type = 'asana'
            elif not source_id and asana_name:
                # Если только название асаны - это название асаны
                object_type = 'asana_name'
            else:
                # По умолчанию - асана
                object_type = 'asana'
        
        db = get_db_session()
        
        # Очищаем фото из import_data перед сохранением, чтобы не засорять БД
        # Base64 фото может быть очень длинным (сотни тысяч символов)
        # Вместо фото сохраняем только информацию о том, что фото было
        if import_data:
            import_data_clean = import_data.copy()
            
            # Обрабатываем одиночное фото
            if 'photo' in import_data_clean:
                photo_data = import_data_clean.get('photo', '')
                if isinstance(photo_data, str) and photo_data.startswith('data:'):
                    # Проверяем, валидное ли фото
                    if ',' in photo_data:
                        base64_part = photo_data.split(',')[1]
                        # Если фото слишком маленькое или обрезанное - удаляем
                        if len(base64_part) < 100:
                            logger.warning(f"Row {row_number}: Base64 photo data too short ({len(base64_part)} chars), removing from import_data")
                            import_data_clean['photo'] = None
                            import_data_clean['photo_info'] = 'removed_too_short'
                        elif len(base64_part) % 4 != 0:
                            logger.warning(f"Row {row_number}: Base64 photo data length not multiple of 4 ({len(base64_part)}), may be truncated, removing")
                            import_data_clean['photo'] = None
                            import_data_clean['photo_info'] = 'removed_truncated'
                        else:
                            # Фото валидное, но слишком длинное для БД - заменяем на метаданные
                            photo_size_kb = len(base64_part) / 1024
                            if photo_size_kb > 100:  # Если больше 100KB base64
                                logger.info(f"Row {row_number}: Photo too large ({photo_size_kb:.1f}KB base64), replacing with metadata")
                                import_data_clean['photo'] = None
                                import_data_clean['photo_info'] = f'removed_too_large_{photo_size_kb:.0f}KB'
                            else:
                                # Фото небольшое, оставляем как есть
                                import_data_clean['photo_info'] = f'embedded_{photo_size_kb:.1f}KB'
            
            # Обрабатываем список фото
            if 'photos' in import_data_clean:
                photos_list = import_data_clean.get('photos', [])
                if photos_list:
                    valid_photos = []
                    for i, photo in enumerate(photos_list):
                        if isinstance(photo, str) and photo.startswith('data:'):
                            if ',' in photo:
                                base64_part = photo.split(',')[1]
                                if len(base64_part) >= 100 and len(base64_part) % 4 == 0:
                                    photo_size_kb = len(base64_part) / 1024
                                    if photo_size_kb <= 100:  # Оставляем только небольшие фото
                                        valid_photos.append(photo)
                    
                    if valid_photos:
                        import_data_clean['photos'] = valid_photos
                        import_data_clean['photo_info'] = f'{len(valid_photos)}_photos_kept_{len(photos_list) - len(valid_photos)}_removed'
                    else:
                        import_data_clean['photos'] = None
                        import_data_clean['photo_info'] = f'all_{len(photos_list)}_photos_removed'
            
            import_data = import_data_clean

        if _find_duplicate_unresolved(
            db,
            asana_name or "",
            source_id,
            error_message,
            moderation_type,
            object_type,
            suggested_name_ru,
            suggested_name_sanskrit,
            suggested_transliteration,
            suggested_definition,
            existing_name_id,
            existing_name_ru,
        ):
            db.close()
            return

        moderation_item = ModerationItem(
            asana_name=asana_name,
            source_id=source_id,
            error_message=error_message,
            row_number=row_number,
            import_data=json.dumps(import_data, ensure_ascii=False) if import_data else None,
            created_at=datetime.now().isoformat(),
            resolved=False,
            moderation_type=moderation_type,
            suggested_name_ru=suggested_name_ru,
            suggested_name_sanskrit=suggested_name_sanskrit,
            suggested_transliteration=suggested_transliteration,
            suggested_definition=suggested_definition,
            existing_name_id=existing_name_id,
            existing_name_ru=existing_name_ru,
            object_type=object_type
        )
        db.add(moderation_item)
        db.commit()
        db.close()
    except Exception as e:
        logger.error(f"Failed to save moderation item: {e}")


def run_asanas_indexed_rows(
    indexed_rows: List[Tuple[int, Dict[str, Any]]],
    source_id: str,
    user: Optional[str] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Dict[str, Any]:
    """
    Импорт асан по уже нормализованным строкам (idx = номер строки Excel).
    Используется напрямую и после чтения из import_staging_rows.
    """
    total_rows = len(indexed_rows)
    imported = 0
    errors = []

    for idx, normalized in indexed_rows:
        try:
            
            if not normalized.get('name_ru'):
                error_msg = f"Строка {idx}: отсутствует название асаны"
                errors.append(error_msg)
                save_moderation_item(
                    asana_name="",
                    source_id=source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized,
                    user=user
                )
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue
            
            name_ru = str(normalized['name_ru']).strip()
            if not name_ru:
                error_msg = f"Строка {idx}: пустое название асаны"
                errors.append(error_msg)
                save_moderation_item(
                    asana_name="",
                    source_id=source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized,
                    user=user
                )
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue
            
            # Ищем существующее название асаны (только точное совпадение 100%, без учета регистра)
            name_id, _ = find_matching_asana_name(name_ru, fuzzy_threshold=1.0)
            
            if not name_id:
                # Если не найдено точное совпадение, отправляем в модерацию
                error_msg = f"Строка {idx}: название '{name_ru}' не найдено в существующих (требуется 100% совпадение)"
                errors.append(error_msg)
                save_moderation_item(
                    asana_name=name_ru,
                    source_id=source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized,
                    user=user,
                    moderation_type="name_mismatch",
                    suggested_name_ru=normalized.get('name_ru')
                )
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue  # Пропускаем эту строку
            
            # Проверяем, существует ли уже асана с таким названием и источником
            from app.ontology import find_existing_asana
            existing_asana_id = find_existing_asana(name_id, source_id)
            
            if existing_asana_id:
                # Асана уже существует, проверяем фото
                from rdflib import Graph, URIRef
                from app.ontology import get_graph, ASANA
                g = get_graph()
                asana_uri = URIRef(existing_asana_id)
                source_uri = URIRef(source_id)
                
                # Получаем все фото этой асаны с этим источником
                existing_photos = list(g.objects(asana_uri, ASANA.hasPhoto))
                existing_photo_hashes = set()
                existing_photo_paths = set()  # Для обратной совместимости
                for existing_photo_uri in existing_photos:
                    existing_photo_source = g.value(existing_photo_uri, ASANA.hasSource)
                    if existing_photo_source == source_uri:
                        # Приоритет: проверяем хеш, если есть
                        existing_hash = g.value(existing_photo_uri, ASANA.photoHash)
                        if existing_hash:
                            existing_photo_hashes.add(str(existing_hash))
                        # Для обратной совместимости также сохраняем пути
                        existing_s3_path = g.value(existing_photo_uri, ASANA.s3PhotoPath)
                        if existing_s3_path:
                            existing_photo_paths.add(str(existing_s3_path))
                
                # Обрабатываем фото из импорта
                photo_s3_paths = []
                photo_hashes = []
                photos_list = normalized.get('photos') or []
                if normalized.get('photo') and not photos_list:
                    photos_list = [normalized.get('photo')]
                
                for photo_base64 in photos_list:
                    if not photo_base64:
                        continue
                    try:
                        # Преобразуем в base64 если нужно
                        if isinstance(photo_base64, bytes):
                            photo_base64 = base64.b64encode(photo_base64).decode('utf-8')
                        
                        # Если это строка base64 без префикса data:, добавляем префикс
                        photo_base64_str = str(photo_base64).strip()
                        photo_base64_str = photo_base64_str.replace('\n', '').replace('\r', '').replace(' ', '')
                        
                        if not photo_base64_str.startswith('data:'):
                            # Определяем формат изображения по содержимому base64
                            try:
                                decoded = base64.b64decode(photo_base64_str, validate=True)
                                if decoded.startswith(b'\x89PNG'):
                                    mime_type = 'image/png'
                                elif decoded.startswith(b'\xff\xd8\xff'):
                                    mime_type = 'image/jpeg'
                                elif decoded.startswith(b'GIF8'):
                                    mime_type = 'image/gif'
                                elif decoded.startswith(b'RIFF') and b'WEBP' in decoded[:12]:
                                    mime_type = 'image/webp'
                                else:
                                    mime_type = 'image/png'
                            except Exception:
                                mime_type = 'image/png'
                            
                            photo_base64_str = f"data:{mime_type};base64,{photo_base64_str}"
                        
                        # Вычисляем хеш ДО загрузки в S3
                        from app.s3_utils import compute_image_hash
                        photo_hash = compute_image_hash(photo_base64_str)
                        
                        # Проверяем, не является ли это фото идентичным уже существующему (по хешу)
                        if photo_hash not in existing_photo_hashes:
                            # Загружаем в S3 только если фото новое
                            photo_s3_path, _ = upload_image_to_s3(photo_base64_str, prefix="asans")
                            photo_s3_paths.append(photo_s3_path)
                            photo_hashes.append(photo_hash)
                        else:
                            logger.info(f"Строка {idx}: фото уже существует для асаны '{name_ru}' (по хешу), пропускаем загрузку в S3")
                    except Exception as e:
                        error_msg = f"Строка {idx}: ошибка загрузки фото в S3: {str(e)}"
                        logger.warning(error_msg)
                        continue
                
                # Проверяем, есть ли новые фото для добавления
                if photo_s3_paths:
                    # Есть новые фото, добавляем их к существующей асане
                    add_asana(name_id, source_id, photo_s3_paths, photo_hashes)
                    logger.info(f"Строка {idx}: добавлено новое фото к существующей асане '{name_ru}'")
                elif photos_list:
                    # Все фото были дубликатами - идентичная запись, пропускаем
                    logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует с идентичными фото, пропускаем")
                else:
                    # Фото нет в импорте
                    if existing_photo_paths:
                        # У асаны есть фото, в импорте нет - пропускаем
                        logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует с фото, пропускаем (в импорте фото нет)")
                    else:
                        # У асаны нет фото и в импорте нет - идентичная запись, пропускаем
                        logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует без фото, пропускаем (идентичная запись)")
            else:
                # Асана не существует, обрабатываем фото и создаем новую
                photo_s3_paths = []
                photo_hashes = []
                photos_list = normalized.get('photos') or []
                if normalized.get('photo') and not photos_list:
                    photos_list = [normalized.get('photo')]
                
                for photo_base64 in photos_list:
                    if not photo_base64:
                        continue
                    try:
                        # Преобразуем в base64 если нужно
                        if isinstance(photo_base64, bytes):
                            photo_base64 = base64.b64encode(photo_base64).decode('utf-8')
                        
                        # Если это строка base64 без префикса data:, добавляем префикс
                        photo_base64_str = str(photo_base64).strip()
                        photo_base64_str = photo_base64_str.replace('\n', '').replace('\r', '').replace(' ', '')
                        
                        if not photo_base64_str.startswith('data:'):
                            # Определяем формат изображения по содержимому base64
                            try:
                                decoded = base64.b64decode(photo_base64_str, validate=True)
                                if decoded.startswith(b'\x89PNG'):
                                    mime_type = 'image/png'
                                elif decoded.startswith(b'\xff\xd8\xff'):
                                    mime_type = 'image/jpeg'
                                elif decoded.startswith(b'GIF8'):
                                    mime_type = 'image/gif'
                                elif decoded.startswith(b'RIFF') and b'WEBP' in decoded[:12]:
                                    mime_type = 'image/webp'
                                else:
                                    mime_type = 'image/png'
                            except Exception:
                                mime_type = 'image/png'
                            
                            photo_base64_str = f"data:{mime_type};base64,{photo_base64_str}"
                        
                        # Загружаем в S3 (возвращает кортеж (путь, хеш))
                        photo_s3_path, photo_hash = upload_image_to_s3(photo_base64_str, prefix="asans")
                        photo_s3_paths.append(photo_s3_path)
                        photo_hashes.append(photo_hash)
                    except Exception as e:
                        error_msg = f"Строка {idx}: ошибка загрузки фото в S3: {str(e)}"
                        logger.warning(error_msg)
                        continue
                
                # Создаем новую асану
                add_asana(name_id, source_id, photo_s3_paths, photo_hashes)
                imported += 1
            
            # Обновляем прогресс
            if progress_callback:
                progress_callback(idx - 1, total_rows)
            
        except Exception as e:
            error_msg = f"Строка {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
            # Сохраняем в модерацию
            try:
                asana_name = normalized.get('name_ru', '') if 'normalized' in locals() else ''
                save_moderation_item(
                    asana_name=str(asana_name),
                    source_id=source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized if 'normalized' in locals() else None,
                    user=user
                )
            except Exception as save_error:
                logger.error(f"Failed to save moderation item: {save_error}")
            
            # Обновляем прогресс даже при ошибке
            if progress_callback:
                progress_callback(idx - 1, total_rows)
    
    return {
        "imported": imported,
        "errors": errors
    }


def import_asanas_from_excel(
    file_path: str,
    source_id: str,
    user: Optional[str] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Dict[str, Any]:
    """
    Импортирует асаны из Excel файла с привязкой к существующему источнику.
    """
    rows = parse_excel_file(file_path)
    indexed = [(idx, normalize_column_names(row)) for idx, row in enumerate(rows, start=2)]
    return run_asanas_indexed_rows(indexed, source_id, user=user, progress_callback=progress_callback)


def scan_sources_from_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Сканирует Excel файл и возвращает список уникальных источников из файла.
    Используется для предварительного просмотра перед импортом.
    
    Returns:
        List[Dict]: Список словарей с данными источников и информацией о том, существуют ли они
    """
    rows = parse_excel_file(file_path)
    sources_found = {}  # {(title, author, year): source_data}
    
    for row in rows:
        normalized = normalize_column_names(row)
        
        if has_source_data(normalized):
            source_data = {
                "title": normalized.get('source_title', 'Неизвестный источник'),
                "author": normalized.get('source_author', ''),
                "year": int(normalized.get('source_year', 0)) if normalized.get('source_year') else 0,
            }
            
            if normalized.get('source_publisher'):
                source_data["publisher"] = normalized['source_publisher']
            if normalized.get('source_pages'):
                source_data["pages"] = int(normalized['source_pages'])
            if normalized.get('source_annotation'):
                source_data["annotation"] = normalized['source_annotation']
            
            # Используем ключ для уникальности
            cache_key = (
                source_data.get('title', '').strip().lower(),
                source_data.get('author', '').strip().lower(),
                source_data.get('year', 0)
            )
            
            if cache_key not in sources_found:
                # Проверяем, существует ли уже такой источник
                existing_source_id = find_existing_source(source_data)
                source_data['exists'] = existing_source_id is not None
                source_data['existing_id'] = existing_source_id
                sources_found[cache_key] = source_data
    
    return list(sources_found.values())


def run_full_indexed_rows(
    indexed_rows: List[Tuple[int, Dict[str, Any]]],
    user: Optional[str] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    source_mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Полный импорт по списку (номер_строки, нормализованный dict).
    """
    total_rows = len(indexed_rows)
    imported_asanas = 0
    imported_sources = 0
    errors = []
    current_source_id = None

    source_cache = {}  # {tuple(title, author, year): source_id}

    for idx, normalized in indexed_rows:
        try:
            
            # Проверяем, нужно ли создать новый источник
            if has_source_data(normalized):
                source_data = {
                    "title": normalized.get('source_title', 'Неизвестный источник'),
                    "author": normalized.get('source_author', ''),
                    "year": int(normalized.get('source_year', 0)) if normalized.get('source_year') else 0,
                }
                
                if normalized.get('source_publisher'):
                    source_data["publisher"] = normalized['source_publisher']
                if normalized.get('source_pages'):
                    source_data["pages"] = int(normalized['source_pages'])
                if normalized.get('source_annotation'):
                    source_data["annotation"] = normalized['source_annotation']
                
                # Проверяем кеш
                cache_key = (
                    source_data.get('title', '').strip().lower(),
                    source_data.get('author', '').strip().lower(),
                    source_data.get('year', 0)
                )
                
                if cache_key in source_cache:
                    # Используем из кеша
                    current_source_id = source_cache[cache_key]
                else:
                    # Проверяем, существует ли уже такой источник
                    existing_source_id = find_existing_source(source_data)
                    
                    if existing_source_id:
                        # Используем существующий источник
                        current_source_id = existing_source_id
                        source_cache[cache_key] = current_source_id  # Сохраняем в кеш
                    else:
                        # Проверяем маппинг источников
                        source_key = f"{source_data.get('title', '')}|{source_data.get('author', '')}|{source_data.get('year', 0)}"
                        if source_mapping and source_key in source_mapping:
                            mapping_value = source_mapping[source_key]
                            # Если значение 'new', создаем новый источник
                            if mapping_value == 'new':
                                current_source_id = add_source(source_data, check_existing=False)
                                source_cache[cache_key] = current_source_id
                                imported_sources += 1
                            else:
                                # Используем существующий источник из маппинга (ID источника)
                                current_source_id = mapping_value
                                source_cache[cache_key] = current_source_id
                        else:
                            # Создаем новый источник (если маппинга нет или ключа нет в маппинге)
                            current_source_id = add_source(source_data, check_existing=False)
                            source_cache[cache_key] = current_source_id  # Сохраняем в кеш
                            imported_sources += 1
            
            # Если нет текущего источника, пропускаем
            if not current_source_id:
                errors.append(f"Строка {idx}: нет источника для асаны")
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue
            
            # Проверяем обязательное поле для асаны
            name_ru = normalized.get('name_ru')
            if not name_ru:
                error_msg = f"Строка {idx}: отсутствует название асаны"
                errors.append(error_msg)
                logger.warning(error_msg)
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue  # Пропускаем строки без названия асаны
            
            name_ru = str(name_ru).strip()
            if not name_ru:
                error_msg = f"Строка {idx}: пустое название асаны"
                errors.append(error_msg)
                logger.warning(error_msg)
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue
            
            # Ищем существующее название асаны (только точное совпадение 100%, без учета регистра)
            name_id, _ = find_matching_asana_name(name_ru, fuzzy_threshold=1.0)
            
            if not name_id:
                # Если не найдено точное совпадение, отправляем в модерацию
                error_msg = f"Строка {idx}: название '{name_ru}' не найдено в существующих (требуется 100% совпадение)"
                errors.append(error_msg)
                logger.warning(error_msg)
                save_moderation_item(
                    asana_name=name_ru,
                    source_id=current_source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized,
                    user=user,
                    moderation_type="name_mismatch",
                    suggested_name_ru=normalized.get('name_ru')
                )
                if progress_callback:
                    progress_callback(idx - 1, total_rows)
                continue  # Пропускаем эту строку
            
            # Проверяем, существует ли уже асана с таким названием и источником
            from app.ontology import find_existing_asana
            existing_asana_id = find_existing_asana(name_id, current_source_id)
            
            if existing_asana_id:
                # Асана уже существует, проверяем фото
                from rdflib import Graph, URIRef
                from app.ontology import get_graph, ASANA
                g = get_graph()
                asana_uri = URIRef(existing_asana_id)
                source_uri = URIRef(current_source_id)
                
                # Получаем все фото этой асаны с этим источником
                existing_photos = list(g.objects(asana_uri, ASANA.hasPhoto))
                existing_photo_hashes = set()
                existing_photo_paths = set()  # Для обратной совместимости
                for existing_photo_uri in existing_photos:
                    existing_photo_source = g.value(existing_photo_uri, ASANA.hasSource)
                    if existing_photo_source == source_uri:
                        # Приоритет: проверяем хеш, если есть
                        existing_hash = g.value(existing_photo_uri, ASANA.photoHash)
                        if existing_hash:
                            existing_photo_hashes.add(str(existing_hash))
                        # Для обратной совместимости также сохраняем пути
                        existing_s3_path = g.value(existing_photo_uri, ASANA.s3PhotoPath)
                        if existing_s3_path:
                            existing_photo_paths.add(str(existing_s3_path))
                
                # Обрабатываем фото из импорта
                photo_s3_path = None
                photo_hash = None
                photo_base64 = normalized.get('photo')
                
                if photo_base64:
                    try:
                        # Преобразуем в base64 если нужно
                        if isinstance(photo_base64, bytes):
                            photo_base64 = base64.b64encode(photo_base64).decode('utf-8')
                        
                        # Если это строка base64 без префикса data:, добавляем префикс
                        photo_base64_str = str(photo_base64).strip()
                        photo_base64_str = photo_base64_str.replace('\n', '').replace('\r', '').replace(' ', '')
                        
                        if not photo_base64_str.startswith('data:'):
                            # Определяем формат изображения по содержимому base64
                            try:
                                decoded = base64.b64decode(photo_base64_str, validate=True)
                                if decoded.startswith(b'\x89PNG'):
                                    mime_type = 'image/png'
                                elif decoded.startswith(b'\xff\xd8\xff'):
                                    mime_type = 'image/jpeg'
                                elif decoded.startswith(b'GIF8'):
                                    mime_type = 'image/gif'
                                elif decoded.startswith(b'RIFF') and b'WEBP' in decoded[:12]:
                                    mime_type = 'image/webp'
                                else:
                                    mime_type = 'image/png'
                            except Exception:
                                mime_type = 'image/png'
                            
                            photo_base64_str = f"data:{mime_type};base64,{photo_base64_str}"
                        
                        # Вычисляем хеш ДО загрузки в S3
                        from app.s3_utils import compute_image_hash
                        photo_hash = compute_image_hash(photo_base64_str)
                        
                        # Проверяем, не является ли это фото идентичным уже существующему (по хешу)
                        if photo_hash and photo_hash in existing_photo_hashes:
                            logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует с идентичным фото (по хешу), пропускаем загрузку в S3")
                            if progress_callback:
                                progress_callback(idx - 1, total_rows)
                            continue  # Пропускаем идентичную запись, НЕ загружаем в S3
                        
                        # Загружаем в S3 только если фото новое
                        photo_s3_path, _ = upload_image_to_s3(photo_base64_str, prefix="asans")
                        
                    except Exception as e:
                        error_msg = f"Строка {idx}: ошибка загрузки фото в S3: {str(e)}"
                        logger.error(error_msg)
                        logger.warning(error_msg)
                        save_moderation_item(
                            asana_name=name_ru,
                            source_id=current_source_id,
                            error_message=error_msg,
                            row_number=idx,
                            import_data=normalized,
                            user=user
                        )
                        photo_s3_path = None
                        photo_hash = None
                else:
                    # Если фото нет в импорте, проверяем есть ли у асаны фото
                    if existing_photo_paths:
                        # У асаны уже есть фото, а в импорте нет - это не идентичная запись, но пропускаем
                        logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует с фото, пропускаем (в импорте фото нет)")
                        if progress_callback:
                            progress_callback(idx - 1, total_rows)
                        continue
                    else:
                        # У асаны нет фото и в импорте нет - идентичная запись, пропускаем
                        logger.info(f"Строка {idx}: асана '{name_ru}' с источником уже существует без фото, пропускаем (идентичная запись)")
                        if progress_callback:
                            progress_callback(idx - 1, total_rows)
                        continue
                
                # Если дошли сюда - фото новое, добавляем его к существующей асане
                photo_paths = [photo_s3_path] if photo_s3_path else []
                photo_hashes_list = [photo_hash] if photo_hash else []
                asana_id = add_asana(name_id, current_source_id, photo_paths, photo_hashes_list)
                logger.info(f"Строка {idx}: добавлено новое фото к существующей асане '{name_ru}'")
            else:
                # Асана не существует, обрабатываем фото и создаем новую
                photo_s3_path = None
                photo_hash = None
                photo_base64 = normalized.get('photo')
                
                if photo_base64:
                    try:
                        # Преобразуем в base64 если нужно
                        if isinstance(photo_base64, bytes):
                            photo_base64 = base64.b64encode(photo_base64).decode('utf-8')
                        
                        # Если это строка base64 без префикса data:, добавляем префикс
                        photo_base64_str = str(photo_base64).strip()
                        photo_base64_str = photo_base64_str.replace('\n', '').replace('\r', '').replace(' ', '')
                        
                        if not photo_base64_str.startswith('data:'):
                            # Определяем формат изображения по содержимому base64
                            try:
                                decoded = base64.b64decode(photo_base64_str, validate=True)
                                if decoded.startswith(b'\x89PNG'):
                                    mime_type = 'image/png'
                                elif decoded.startswith(b'\xff\xd8\xff'):
                                    mime_type = 'image/jpeg'
                                elif decoded.startswith(b'GIF8'):
                                    mime_type = 'image/gif'
                                elif decoded.startswith(b'RIFF') and b'WEBP' in decoded[:12]:
                                    mime_type = 'image/webp'
                                else:
                                    mime_type = 'image/png'
                            except Exception:
                                mime_type = 'image/png'
                            
                            photo_base64_str = f"data:{mime_type};base64,{photo_base64_str}"
                        
                        # Загружаем в S3 (возвращает кортеж (путь, хеш))
                        photo_s3_path, photo_hash = upload_image_to_s3(photo_base64_str, prefix="asans")
                        
                    except Exception as e:
                        error_msg = f"Строка {idx}: ошибка загрузки фото в S3: {str(e)}"
                        logger.error(error_msg)
                        logger.warning(error_msg)
                        save_moderation_item(
                            asana_name=name_ru,
                            source_id=current_source_id,
                            error_message=error_msg,
                            row_number=idx,
                            import_data=normalized,
                            user=user
                        )
                        photo_s3_path = None
                        photo_hash = None
                        photo_hash = None
                
                # Создаем новую асану
                photo_paths = [photo_s3_path] if photo_s3_path else []
                photo_hashes_list = [photo_hash] if photo_hash else []
                asana_id = add_asana(name_id, current_source_id, photo_paths, photo_hashes_list)
                imported_asanas += 1
            
            # Обновляем прогресс
            if progress_callback:
                progress_callback(idx - 1, total_rows)
            
        except Exception as e:
            error_msg = f"Строка {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)
            # Сохраняем в модерацию
            try:
                asana_name = normalized.get('name_ru', '') if 'normalized' in locals() else ''
                save_moderation_item(
                    asana_name=str(asana_name),
                    source_id=current_source_id,
                    error_message=error_msg,
                    row_number=idx,
                    import_data=normalized if 'normalized' in locals() else None,
                    user=user
                )
            except Exception as save_error:
                logger.error(f"Failed to save moderation item: {save_error}")
            
            # Обновляем прогресс даже при ошибке
            if progress_callback:
                progress_callback(idx - 1, total_rows)
    
    return {
        "imported_asanas": imported_asanas,
        "imported_sources": imported_sources,
        "errors": errors
    }


def import_full_from_excel(
    file_path: str,
    user: Optional[str] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    source_mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Импортирует асаны и источники из Excel файла (парсинг файла → run_full_indexed_rows).
    """
    rows = parse_excel_file(file_path)
    indexed = [(idx, normalize_column_names(row)) for idx, row in enumerate(rows, start=2)]
    return run_full_indexed_rows(
        indexed, user=user, progress_callback=progress_callback, source_mapping=source_mapping
    )


def run_asana_names_indexed_rows(
    indexed_rows: List[Tuple[int, Dict[str, Any]]],
    user: Optional[str] = None,
) -> Dict[str, Any]:
    """Импорт названий по списку (номер строки, нормализованные поля)."""
    imported = 0
    skipped = 0
    skipped_items: List[Dict[str, Any]] = []
    errors = []

    for idx, normalized in indexed_rows:
        try:
            
            if not normalized.get('name_ru'):
                error_msg = f"Строка {idx}: отсутствует название асаны"
                errors.append(error_msg)
                continue
            
            name_ru = str(normalized['name_ru']).strip()
            if not name_ru:
                error_msg = f"Строка {idx}: пустое название асаны"
                errors.append(error_msg)
                continue
            
            # Проверяем, существует ли уже такое название (только точное совпадение 100%)
            existing_name_id, _ = find_matching_asana_name(name_ru, fuzzy_threshold=1.0)
            
            if existing_name_id:
                skipped += 1
                skipped_items.append({"row": idx, "name": name_ru})
                continue
            
            # Создаем новое название асаны
            name_data = {"name_ru": name_ru}
            
            if normalized.get('name_sanskrit'):
                name_data["name_sanskrit"] = str(normalized['name_sanskrit']).strip()
            if normalized.get('transliteration'):
                name_data["transliteration"] = str(normalized['transliteration']).strip()
            if normalized.get('definition'):
                name_data["definition"] = str(normalized['definition']).strip()
            
            add_asana_name(name_data)
            imported += 1

        except Exception as e:
            error_msg = f"Строка {idx}: {str(e)}"
            errors.append(error_msg)
            logger.error(error_msg)

    return {
        "imported": imported,
        "skipped": skipped,
        "skipped_items": skipped_items,
        "errors": errors,
    }


def import_asana_names_from_excel(file_path: str, user: Optional[str] = None) -> Dict[str, Any]:
    """
    Импортирует названия асан из Excel файла.
    """
    rows = parse_excel_file(file_path)
    indexed = [(idx, normalize_column_names(row)) for idx, row in enumerate(rows, start=2)]
    return run_asana_names_indexed_rows(indexed, user=user)


def export_moderation_to_excel(items: List[Dict[str, Any]]) -> BytesIO:
    """
    Экспортирует записи модерации в Excel файл в том же формате, что и при импорте.
    Изображения вставляются как картинки в Excel.
    
    Args:
        items: Список записей модерации с полями:
            - row_number: номер строки
            - import_data: словарь с данными импорта (может содержать photo, photo_base64, photo_url)
            - asana_name: название асаны
            - error_message: сообщение об ошибке
            - и другие поля
    
    Returns:
        BytesIO: поток с Excel файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Модерация"
    
    # Определяем колонки на основе данных импорта
    # Используем только те поля, которые реально используются при импорте асан
    columns = [
        'название',  # name_ru
        'фото',  # photo (будет изображение)
        'название источника',  # source_title
        'автор',  # source_author
        'год',  # source_year
        'издательство',  # source_publisher
        'страницы',  # source_pages
        'аннотация',  # source_annotation
    ]
    
    # Заголовки
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
    
    # Заполняем данные
    for item_idx, item in enumerate(items, start=2):
        import_data = item.get('import_data', {}) or {}
        
        # Название асаны
        ws.cell(row=item_idx, column=1, value=item.get('asana_name') or import_data.get('name_ru', ''))
        
        # Фото - вставляем как изображение
        photo_col = 2
        photo_data = None
        
        # Пробуем получить фото из разных источников
        if import_data.get('photo_base64'):
            photo_data = import_data['photo_base64']
        elif import_data.get('photo'):
            photo_data = import_data['photo']
        elif import_data.get('photo_url'):
            # Если это URL, пропускаем (не загружаем из интернета при экспорте)
            # Можно было бы загрузить, но это может быть медленно и требует requests
            photo_data = None
        
        # Вставляем изображение в Excel
        if photo_data:
            try:
                # Декодируем base64
                if isinstance(photo_data, str):
                    # Убираем префикс data: если есть
                    if photo_data.startswith('data:'):
                        photo_data = photo_data.split(',', 1)[1]
                    
                    # Декодируем base64
                    try:
                        img_bytes = base64.b64decode(photo_data)
                    except Exception:
                        # Если не base64, пробуем как URL или путь
                        img_bytes = None
                else:
                    img_bytes = photo_data
                
                if img_bytes:
                    # Создаем временный файл изображения
                    img_io = BytesIO(img_bytes)
                    
                    # Создаем объект Image
                    img = Image(img_io)
                    
                    # Устанавливаем размер (опционально, можно настроить)
                    img.width = 200
                    img.height = 200
                    
                    # Вставляем изображение в ячейку
                    cell_ref = ws.cell(row=item_idx, column=photo_col).coordinate
                    ws.add_image(img, cell_ref)
                    
                    # Увеличиваем высоту строки для изображения
                    ws.row_dimensions[item_idx].height = 150
            except Exception as e:
                logger.warning(f"Не удалось вставить изображение в строку {item_idx}: {e}")
        
        # Данные источника
        ws.cell(row=item_idx, column=3, value=import_data.get('source_title', ''))
        ws.cell(row=item_idx, column=4, value=import_data.get('source_author', ''))
        ws.cell(row=item_idx, column=5, value=import_data.get('source_year', ''))
        ws.cell(row=item_idx, column=6, value=import_data.get('source_publisher', ''))
        ws.cell(row=item_idx, column=7, value=import_data.get('source_pages', ''))
        ws.cell(row=item_idx, column=8, value=import_data.get('source_annotation', ''))
    
    # Настраиваем ширину колонок
    column_widths = {
        1: 30,  # название
        2: 25,  # фото
        3: 30,  # название источника
        4: 25,  # автор
        5: 10,  # год
        6: 25,  # издательство
        7: 15,  # страницы
        8: 40,  # аннотация
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_asana_names_to_excel() -> BytesIO:
    """
    Выгружает названия асан из онтологии в Excel в том же формате колонок,
    что ожидает import_asana_names_from_excel (название, санскрит, транслитерация, определение).
    """
    names = load_asana_names()
    wb = Workbook()
    ws = wb.active
    ws.title = "Названия"
    headers = ["название", "санскрит", "транслитерация", "определение"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
    sorted_names = sorted(names, key=lambda x: (x.get("name_ru") or "").lower())
    for row_idx, n in enumerate(sorted_names, start=2):
        ws.cell(row=row_idx, column=1, value=n.get("name_ru") or "")
        ws.cell(row=row_idx, column=2, value=n.get("name_sanskrit") or "")
        ws.cell(row=row_idx, column=3, value=n.get("transliteration") or "")
        ws.cell(row=row_idx, column=4, value=n.get("definition") or "")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

